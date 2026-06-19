# -*- coding: utf-8 -*-
"""
智能仓储 WMS 系统 - 报告与统计导出模块
每周一生成发布统计、趋势图表 PDF/Excel
支持按时间、仓库、版本查询导出
"""

import os
import csv
import json
from datetime import datetime, timedelta
from typing import List, Dict, Optional

from config import (
    REPORT_DIR, EXPORT_DIR, WAREHOUSES, RELEASE_STATUS, RISK_LEVELS,
    get_current_time_str, get_timestamp
)
from logger import log
from database import db


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ReportEngine:
    """报告与统计引擎"""

    def __init__(self):
        os.makedirs(REPORT_DIR, exist_ok=True)
        os.makedirs(EXPORT_DIR, exist_ok=True)

    def _get_week_range(self, ref_date: datetime = None) -> tuple:
        """获取本周一到下周一的时间范围"""
        if not ref_date:
            ref_date = datetime.now()
        monday = ref_date - timedelta(days=ref_date.weekday())
        next_monday = monday + timedelta(days=7)
        start = monday.strftime('%Y-%m-%d 00:00:00')
        end = next_monday.strftime('%Y-%m-%d 00:00:00')
        return start, end, monday.strftime('%Y%m%d')

    def generate_weekly_report(self, ref_date: datetime = None) -> Dict:
        """生成周统计报告（Excel + PDF + TXT）"""
        start, end, week_tag = self._get_week_range(ref_date)
        log.info(f"生成周统计报告: {start} ~ {end}")

        stats = db.get_weekly_stats(start, end)
        releases = db.list_releases(start_time=start, end_time=end, limit=1000)
        rollbacks = db.get_rollback_records(start_time=start, end_time=end)
        drills = db.list_drills(start_time=start, end_time=end, limit=100)

        report_data = {
            'report_title': 'WMS 系统发布周度统计报告',
            'week_range': f"{start.split()[0]} ~ {end.split()[0]}",
            'generated_at': get_current_time_str(),
            'summary': {
                'total_releases': stats.get('total_releases', 0),
                'success_count': stats.get('success_count', 0),
                'rollback_count': stats.get('rollback_count', 0),
                'rejected_count': stats.get('rejected_count', 0),
                'precheck_failed_count': stats.get('precheck_failed_count', 0),
                'normal_count': stats.get('normal_count', 0),
                'emergency_count': stats.get('emergency_count', 0),
                'success_rate': (
                    round(stats.get('success_count', 0) / stats.get('total_releases', 1) * 100, 2)
                    if stats.get('total_releases', 0) > 0 else 0
                ),
                'rollback_rate': (
                    round(stats.get('rollback_count', 0) / stats.get('total_releases', 1) * 100, 2)
                    if stats.get('total_releases', 0) > 0 else 0
                ),
            },
            'status_breakdown': stats.get('status_breakdown', {}),
            'daily_trend': stats.get('daily_trend', []),
            'releases': releases,
            'rollbacks': rollbacks,
            'drills': drills,
        }

        outputs = {}
        outputs['txt'] = self._save_txt_report(report_data, week_tag)
        if HAS_OPENPYXL:
            outputs['xlsx'] = self._save_excel_report(report_data, week_tag)
        if HAS_REPORTLAB:
            outputs['pdf'] = self._save_pdf_report(report_data, week_tag)

        outputs['json'] = self._save_json_report(report_data, week_tag)

        log.info(f"周报告生成完成: {list(outputs.keys())}")
        return {
            'success': True,
            'week_range': report_data['week_range'],
            'files': outputs,
            'summary': report_data['summary'],
        }

    def _save_txt_report(self, data: Dict, week_tag: str) -> str:
        path = os.path.join(REPORT_DIR, f"weekly_report_{week_tag}.txt")
        s = data['summary']
        lines = []
        lines.append("=" * 70)
        lines.append(f"  {data['report_title']}")
        lines.append("=" * 70)
        lines.append(f"统计周期: {data['week_range']}")
        lines.append(f"生成时间: {data['generated_at']}")
        lines.append("")
        lines.append("-" * 50)
        lines.append("【核心指标】")
        lines.append(f"  总发布数:       {s['total_releases']}")
        lines.append(f"  发布成功数:     {s['success_count']}")
        lines.append(f"  回滚次数:       {s['rollback_count']}")
        lines.append(f"  审批拒绝数:     {s['rejected_count']}")
        lines.append(f"  前置检查失败:   {s['precheck_failed_count']}")
        lines.append(f"  成功率:         {s['success_rate']:.2f}%")
        lines.append(f"  回滚率:         {s['rollback_rate']:.2f}%")
        lines.append(f"  常规发布数:     {s['normal_count']}")
        lines.append(f"  紧急发布数:     {s['emergency_count']}")
        lines.append("")

        lines.append("-" * 50)
        lines.append("【发布状态分布】")
        for status, cnt in data['status_breakdown'].items():
            status_name = RELEASE_STATUS.get(status, status)
            lines.append(f"  {status_name}: {cnt}")
        lines.append("")

        lines.append("-" * 50)
        lines.append("【每日发布趋势】")
        for d in data['daily_trend']:
            lines.append(f"  {d['day']}: {d['cnt']} 次发布")
        lines.append("")

        lines.append("-" * 50)
        lines.append(f"【发布明细 (共 {len(data['releases'])} 条)】")
        lines.append(f"  {'单号':<18} {'版本':<12} {'风险':<6} {'状态':<10} {'提交人':<10} {'创建时间'}")
        for r in data['releases']:
            lines.append(
                f"  {r['release_no']:<18} {r['version']:<12} "
                f"{RISK_LEVELS.get(r['risk_level'], {}).get('name', r['risk_level']):<6} "
                f"{RELEASE_STATUS.get(r['status'], r['status']):<10} "
                f"{r['submitter']:<10} {r['created_at']}"
            )
        lines.append("")

        if data['rollbacks']:
            lines.append("-" * 50)
            lines.append(f"【回滚记录 (共 {len(data['rollbacks'])} 条)】")
            for rb in data['rollbacks']:
                lines.append(
                    f"  {rb['release_no']} | {rb['trigger_type']} | "
                    f"影响订单 {rb['affected_orders']} | {rb['started_at']}"
                )
            lines.append("")

        if data['drills']:
            lines.append("-" * 50)
            lines.append(f"【回滚演练 (共 {len(data['drills'])} 条)】")
            for dr in data['drills']:
                lines.append(f"  {dr['drill_no']} | {dr['title']} | {dr['result'] or dr['status']}")

        lines.append("\n" + "=" * 70)
        with open(path, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        return path

    def _save_json_report(self, data: Dict, week_tag: str) -> str:
        path = os.path.join(REPORT_DIR, f"weekly_report_{week_tag}.json")
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return path

    def _save_excel_report(self, data: Dict, week_tag: str) -> str:
        path = os.path.join(REPORT_DIR, f"weekly_report_{week_tag}.xlsx")
        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws1 = wb.active
        ws1.title = "汇总"
        s = data['summary']
        ws1.append(["指标", "数值"])
        ws1.append(["总发布数", s['total_releases']])
        ws1.append(["发布成功数", s['success_count']])
        ws1.append(["回滚次数", s['rollback_count']])
        ws1.append(["审批拒绝数", s['rejected_count']])
        ws1.append(["前置检查失败", s['precheck_failed_count']])
        ws1.append(["发布成功率(%)", s['success_rate']])
        ws1.append(["回滚率(%)", s['rollback_rate']])
        ws1.append(["常规发布数", s['normal_count']])
        ws1.append(["紧急发布数", s['emergency_count']])
        for row in ws1.iter_rows(min_row=1, max_row=1, max_col=2):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        ws2 = wb.create_sheet("每日趋势")
        ws2.append(["日期", "发布次数"])
        for d in data['daily_trend']:
            ws2.append([d['day'], d['cnt']])
        for row in ws2.iter_rows(min_row=1, max_row=1, max_col=2):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        if len(data['daily_trend']) > 1:
            chart = LineChart()
            chart.title = "每日发布趋势"
            chart.y_axis.title = "发布次数"
            chart.x_axis.title = "日期"
            data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(data['daily_trend']) + 1)
            cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(data['daily_trend']) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws2.add_chart(chart, "D2")

        ws3 = wb.create_sheet("发布明细")
        headers = ["发布单号", "版本", "风险级别", "状态", "标题", "提交人", "创建时间", "完成时间"]
        ws3.append(headers)
        for r in data['releases']:
            ws3.append([
                r['release_no'], r['version'],
                RISK_LEVELS.get(r['risk_level'], {}).get('name', r['risk_level']),
                RELEASE_STATUS.get(r['status'], r['status']),
                r['title'], r['submitter'], r['created_at'], r.get('finished_at', '')
            ])
        for row in ws3.iter_rows(min_row=1, max_row=1, max_col=len(headers)):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        ws4 = wb.create_sheet("状态分布")
        ws4.append(["状态", "数量"])
        for status, cnt in data['status_breakdown'].items():
            ws4.append([RELEASE_STATUS.get(status, status), cnt])
        for row in ws4.iter_rows(min_row=1, max_row=1, max_col=2):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        if data['status_breakdown']:
            pie = PieChart()
            pie.title = "发布状态分布"
            data_ref = Reference(ws4, min_col=2, min_row=1,
                                 max_row=len(data['status_breakdown']) + 1)
            cats_ref = Reference(ws4, min_col=1, min_row=2,
                                 max_row=len(data['status_breakdown']) + 1)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            ws4.add_chart(pie, "D2")

        ws5 = wb.create_sheet("回滚记录")
        rb_headers = ["发布单号", "触发类型", "触发原因", "回滚版本",
                      "影响订单", "状态", "开始时间", "完成时间"]
        ws5.append(rb_headers)
        for rb in data['rollbacks']:
            ws5.append([
                rb['release_no'], rb['trigger_type'], rb.get('trigger_reason', ''),
                rb['rollback_version'], rb.get('affected_orders', 0),
                rb['status'], rb['started_at'], rb.get('finished_at', '')
            ])
        for row in ws5.iter_rows(min_row=1, max_row=1, max_col=len(rb_headers)):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        for ws in wb.worksheets:
            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        wb.save(path)
        return path

    def _save_pdf_report(self, data: Dict, week_tag: str) -> str:
        path = os.path.join(REPORT_DIR, f"weekly_report_{week_tag}.pdf")
        doc = SimpleDocTemplate(path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=20)
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
        normal_style = styles['Normal']

        elements.append(Paragraph(data['report_title'], title_style))
        elements.append(Paragraph(f"统计周期: {data['week_range']}", normal_style))
        elements.append(Paragraph(f"生成时间: {data['generated_at']}", normal_style))
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("核心指标", h2_style))
        s = data['summary']
        table_data = [
            ['指标', '数值'],
            ['总发布数', str(s['total_releases'])],
            ['发布成功数', str(s['success_count'])],
            ['回滚次数', str(s['rollback_count'])],
            ['审批拒绝数', str(s['rejected_count'])],
            ['发布成功率', f"{s['success_rate']:.2f}%"],
            ['回滚率', f"{s['rollback_rate']:.2f}%"],
            ['常规发布数', str(s['normal_count'])],
            ['紧急发布数', str(s['emergency_count'])],
        ]
        t = Table(table_data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("每日发布趋势", h2_style))
        if data['daily_trend']:
            trend_data = [['日期', '发布次数']]
            for d in data['daily_trend']:
                trend_data.append([d['day'], str(d['cnt'])])
            t2 = Table(trend_data, colWidths=[200, 100])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(t2)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("发布状态分布", h2_style))
        if data['status_breakdown']:
            sb_data = [['状态', '数量']]
            for status, cnt in data['status_breakdown'].items():
                sb_data.append([RELEASE_STATUS.get(status, status), str(cnt)])
            t3 = Table(sb_data, colWidths=[200, 100])
            t3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFC000')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(t3)

        doc.build(elements)
        return path

    def query_and_export(self, start_time: str = None, end_time: str = None,
                         status: str = None, risk_level: str = None,
                         warehouse_id: str = None, version: str = None,
                         export_format: str = 'csv',
                         data_type: str = 'releases') -> str:
        """
        按条件查询并导出
        :param export_format: 'csv' | 'xlsx' | 'txt' | 'json'
        :param data_type: 'releases' | 'rollbacks' | 'monitors' | 'approvals'
        """
        log.info(f"导出查询: 类型={data_type}, 格式={export_format}, "
                 f"时间={start_time}~{end_time}, 仓库={warehouse_id}, 版本={version}")

        release_ids = None
        if data_type in ('rollbacks', 'monitors', 'approvals'):
            if version or status or risk_level:
                filter_releases = db.list_releases(
                    status=status, risk_level=risk_level,
                    start_time=start_time, end_time=end_time,
                    version=version, limit=100000
                )
                release_ids = [r['id'] for r in filter_releases]

        if data_type == 'releases':
            records = db.list_releases(
                status=status, risk_level=risk_level,
                start_time=start_time, end_time=end_time,
                version=version, limit=10000
            )
        elif data_type == 'rollbacks':
            records = db.get_rollback_records(
                start_time=start_time, end_time=end_time
            )
            if release_ids is not None:
                records = [r for r in records if r['release_id'] in release_ids]
            if warehouse_id:
                filtered = []
                for r in records:
                    affected = r.get('affected_warehouses') or ''
                    if isinstance(affected, str):
                        try:
                            affected = json.loads(affected)
                        except:
                            affected = [affected]
                    wh_name = WAREHOUSES.get(warehouse_id, {}).get('name', warehouse_id)
                    if warehouse_id in affected or wh_name in affected:
                        filtered.append(r)
                records = filtered
        elif data_type == 'monitors':
            records = db.get_monitor_records(
                warehouse_id=warehouse_id,
                start_time=start_time, end_time=end_time,
                limit=10000
            )
            if release_ids is not None:
                records = [r for r in records if r['release_id'] in release_ids]
        elif data_type == 'approvals':
            if release_ids is not None:
                records = []
                for rid in release_ids:
                    records.extend(db.get_approvals(rid))
            else:
                releases = db.list_releases(
                    start_time=start_time, end_time=end_time, limit=10000
                )
                records = []
                for r in releases:
                    records.extend(db.get_approvals(r['id']))
        else:
            records = []

        filename = f"export_{data_type}_{get_timestamp()}"
        base_path = os.path.join(EXPORT_DIR, filename)

        if export_format == 'csv':
            return self._export_csv(records, base_path + '.csv', data_type)
        elif export_format == 'xlsx' and HAS_OPENPYXL:
            return self._export_xlsx(records, base_path + '.xlsx', data_type)
        elif export_format == 'json':
            return self._export_json(records, base_path + '.json')
        else:
            return self._export_txt(records, base_path + '.txt', data_type)

    def _export_csv(self, records: List[Dict], path: str, data_type: str) -> str:
        if not records:
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                f.write("No data\n")
            return path
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)
        log.info(f"已导出 CSV: {path} ({len(records)} 条)")
        return path

    def _export_json(self, records: List[Dict], path: str) -> str:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2, default=str)
        log.info(f"已导出 JSON: {path} ({len(records)} 条)")
        return path

    def _export_txt(self, records: List[Dict], path: str, data_type: str) -> str:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(f"Export Report - {data_type}\n")
            f.write(f"Generated: {get_current_time_str()}\n")
            f.write(f"Total Records: {len(records)}\n")
            f.write("=" * 80 + "\n\n")
            for i, r in enumerate(records, 1):
                f.write(f"--- Record {i} ---\n")
                for k, v in r.items():
                    f.write(f"  {k}: {v}\n")
                f.write("\n")
        log.info(f"已导出 TXT: {path} ({len(records)} 条)")
        return path

    def _export_xlsx(self, records: List[Dict], path: str, data_type: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data_type
        if not records:
            ws.append(["No data"])
        else:
            ws.append(list(records[0].keys()))
            for r in records:
                ws.append(list(r.values()))
        wb.save(path)
        log.info(f"已导出 XLSX: {path} ({len(records)} 条)")
        return path


report_engine = ReportEngine()


class WeeklyReportScheduler:
    """周报告定时调度器 - 后台常驻，周一自动生成周报告"""

    def __init__(self):
        self._running = False
        self.pid_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'data', 'weekly_report_scheduler.pid'
        )
        self.last_run_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'data', 'weekly_report_last_run.txt'
        )

    def is_running(self) -> bool:
        if not os.path.exists(self.pid_file):
            return False
        try:
            with open(self.pid_file, 'r') as f:
                pid = int(f.read().strip())
            import ctypes
            kernel32 = ctypes.windll.kernel32
            process = kernel32.OpenProcess(1024, 0, pid)
            if process:
                kernel32.CloseHandle(process)
                return True
            return False
        except Exception:
            return False

    def get_pid(self):
        if os.path.exists(self.pid_file):
            try:
                with open(self.pid_file, 'r') as f:
                    return int(f.read().strip())
            except:
                return None
        return None

    def _get_last_run_date(self):
        if os.path.exists(self.last_run_file):
            try:
                with open(self.last_run_file, 'r') as f:
                    return f.read().strip()
            except:
                return None
        return None

    def _set_last_run_date(self, date_str):
        try:
            with open(self.last_run_file, 'w') as f:
                f.write(date_str)
        except Exception as e:
            log.warning(f"写入周报告上次运行记录失败: {e}")

    def start(self) -> Dict:
        if self.is_running():
            return {'success': False, 'message': '周报告定时任务已在运行', 'pid': self.get_pid()}

        import sys
        import subprocess

        script_path = os.path.abspath(__file__)
        project_dir = os.path.dirname(script_path)
        main_path = os.path.join(project_dir, 'main.py')

        python_exe = sys.executable

        if os.name == 'nt':
            DETACHED_PROCESS = 0x00000008
            CREATE_NEW_PROCESS_GROUP = 0x00000200
            flags = DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP
            subprocess.Popen(
                [python_exe, main_path, 'report', 'scheduler', 'run'],
                cwd=project_dir,
                creationflags=flags,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        else:
            subprocess.Popen(
                [python_exe, main_path, 'report', 'scheduler', 'run'],
                cwd=project_dir,
                start_new_session=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )

        import time
        for _ in range(10):
            time.sleep(0.5)
            if self.is_running():
                break

        pid = self.get_pid()
        log.info(f"周报告定时任务已启动, PID={pid}")
        log.audit("启动周报告定时任务", "system", "weekly_report_scheduler",
                  details={"pid": pid})
        return {'success': True, 'message': '周报告定时任务已启动', 'pid': pid}

    def stop(self) -> Dict:
        if not self.is_running():
            return {'success': True, 'message': '周报告定时任务未在运行'}

        pid = self.get_pid()
        try:
            if os.name == 'nt':
                import ctypes
                kernel32 = ctypes.windll.kernel32
                process = kernel32.OpenProcess(1, 0, pid)
                if process:
                    kernel32.TerminateProcess(process, 0)
                    kernel32.CloseHandle(process)
            else:
                import signal
                os.kill(pid, signal.SIGTERM)
        except Exception as e:
            log.warning(f"停止周报告定时任务异常: {e}")

        try:
            if os.path.exists(self.pid_file):
                os.remove(self.pid_file)
        except:
            pass

        log.info(f"周报告定时任务已停止, PID={pid}")
        log.audit("停止周报告定时任务", "system", "weekly_report_scheduler",
                  details={"pid": pid})
        return {'success': True, 'message': '周报告定时任务已停止', 'pid': pid}

    def status(self) -> Dict:
        running = self.is_running()
        last_run = self._get_last_run_date()
        return {
            'running': running,
            'pid': self.get_pid() if running else None,
            'last_run_date': last_run,
            'schedule': '每周一 09:00 自动生成周报告',
        }

    def run_loop(self):
        pid = os.getpid()
        with open(self.pid_file, 'w') as f:
            f.write(str(pid))

        log.info(f"[周报告调度器] 启动, PID={pid}")
        log.audit("周报告定时任务启动", "system", "weekly_report_scheduler",
                  details={"pid": pid})

        self._running = True
        import signal

        def _handle_signal(signum, frame):
            log.info(f"[周报告调度器] 收到信号 {signum}, 准备退出")
            self._running = False

        if os.name != 'nt':
            signal.signal(signal.SIGTERM, _handle_signal)
            signal.signal(signal.SIGINT, _handle_signal)

        try:
            while self._running:
                try:
                    self._tick()
                except Exception as e:
                    log.error(f"[周报告调度器] 循环异常: {str(e)}", exc_info=True)
                import time
                time.sleep(1800)
        finally:
            try:
                if os.path.exists(self.pid_file):
                    os.remove(self.pid_file)
            except:
                pass
            log.info("[周报告调度器] 已退出")

    def _tick(self):
        now = datetime.now()
        today_str = now.strftime('%Y-%m-%d')
        last_run = self._get_last_run_date()

        if now.weekday() != 0:
            return

        if now.hour < 9:
            return

        if last_run == today_str:
            return

        log.info(f"[周报告调度器] 到点生成周报告: {today_str}")

        try:
            engine = ReportEngine()
            result = engine.generate_weekly_report()
            paths = result.get('paths', {})
            self._set_last_run_date(today_str)
            log.info(f"[周报告调度器] 周报告生成成功: {paths}")
            log.audit("自动生成周报告", "system", "weekly_report",
                      details={
                          "date": today_str,
                          "paths": paths,
                          "generated_at": get_current_time_str(),
                      })
        except Exception as e:
            log.error(f"[周报告调度器] 生成周报告失败: {e}", exc_info=True)
            log.audit("自动生成周报告失败", "system", "weekly_report",
                      details={"date": today_str, "error": str(e)})

    def run_now(self):
        log.info("[周报告调度器] 手动触发周报告生成")
        engine = ReportEngine()
        result = engine.generate_weekly_report()
        paths = result.get('paths', {})
        today_str = datetime.now().strftime('%Y-%m-%d')
        self._set_last_run_date(today_str)
        log.audit("手动触发生成周报告", "system", "weekly_report",
                  details={"date": today_str, "paths": paths})
        return result


weekly_scheduler = WeeklyReportScheduler()
